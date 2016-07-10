# coding=utf-8

from ..ext import mongo
from random import randint
import csv
import os.path
from openpyxl import load_workbook, Workbook
from openpyxl.compat import range as xl_range
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.styles.colors import WHITE
from datetime import datetime, timedelta
import urllib
from ..utils import my_secure_filename


__author__ = 'hany'


class TransformFile(object):
    def __init__(self, file_object):
        self.file = file_object
        self.file_name = my_secure_filename(file_object.filename)
        # print '>????', self.file_name, '<<<', repr(file_object.filename),
        self.content_type = file_object.content_type
        self._allowed = ['text/csv']
        self.ext = os.path.splitext(self.file_name)[1]
        # print '<<<<', self.ext, self.file_name
        self._map = {'.csv': self._transform_csv, '.xls': self._transform_xls,
                     '.xlsx': self._transform_xls}
        self.table_grid = {'define': list(), 'content': list()}

    def transform(self):
        # assert self.content_type in self._allowed, 'Content type cannot be transformed'
        return self._map[self.ext]()

    def _transform_csv(self):
        stored_file = '/tmp/%s.%s' % (self.file_name, randint(0, 100000))
        self.file.save(stored_file)
        csv_reader = csv.reader(open(stored_file, 'rb'))

        i = 0
        for row in csv_reader:
            if i == 0:
                for j in range(len(row)):
                    self.table_grid['define'].append(
                        {'name': 'COL_%s' % j, 'field': 'COL_%s' % j, 'displayName': unicode(row[j], 'utf8'), 'type': 'text'}
                    )
            else:
                tmp_dict = dict()
                for j in range(len(row)):
                    tmp_dict['COL_%s' % j] = unicode(row[j], 'utf8')

                self.table_grid['content'].append(tmp_dict)
            i += 1

        return self

    def _transform_xls(self):
        stored_file = '/tmp/%s.%s' % (randint(0, 100000), self.file_name)
        self.file.save(stored_file)

        work_book = load_workbook(stored_file)
        sheet = work_book[work_book.get_sheet_names()[0]]

        row_id = 0
        for row in sheet.rows:
            if row_id == 0:
                for cell_id in range(len(row)):
                    cell_val = self._xls_cell_value(row[cell_id])
                    self.table_grid['define'].append(
                        {'name': 'COL_%s' % cell_id, 'field': 'COL_%s' % cell_id, 'displayName': cell_val}
                    )
            else:
                tmp_dict = dict()
                for cell_id in range(len(row)):
                    cell_val = self._xls_cell_value(row[cell_id])
                    if cell_val is not None:
                        tmp_dict['COL_%s' % cell_id] = cell_val
                if len(tmp_dict.keys()) > 0:
                    self.table_grid['content'].append(tmp_dict)
            row_id += 1

        return self

    @staticmethod
    def _xls_cell_value(cell):
        cell_type = type(cell.value).__name__
        if cell_type in ('unicode', 'str'):
            if (cell.value[0]) == '=':
                return ''
            else:
                return cell.value
        elif cell_type in ('long', ):
            return cell.value
        elif cell_type in ('datetime', ):
            return cell.value.strftime('%Y-%m-%d %H:%M:%S')
        elif cell_type in ('NoneType', ):
            return None
        else:
            return ''


class XlsData(object):
    def __init__(self, sheet='Sheet1', data=list()):
        self.sheet = sheet
        self.data = data
        self.file_name = '/tmp/%s_%s.xlsx' % (datetime.now().strftime('%Y%m%d%H%M%S'), randint(0,1000))

    def transform_file(self):
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = self.sheet
        for row in xl_range(1, len(self.data)+1):
            for col in xl_range(1, len(self.data[row-1])+1):
                print row, col
                work_sheet.cell(column=col, row=row, value=self.data[row-1][col-1])
        work_sheet.row_dimensions[1].font = Font(color=WHITE, bold=True)
        work_sheet.row_dimensions[1].fill = PatternFill(fill_type='solid', start_color='005C9BD1')
        work_book.save(filename=self.file_name)
        return self
